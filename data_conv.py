from openpyxl import Workbook
from openpyxl.styles.fonts import Font
import json

air = ''
with open('air_filters.json', 'r', encoding='utf-8') as f:
     air = f.read()

oil = ''

with open('oil_filters.json', 'r', encoding='utf-8') as f:
     oil = f.read()


air = json.loads(air)
header = list(air[0].keys())
grand_air = [list(item.values()) for item in air]
grand_air.insert(0, header)

wb = Workbook()
ws = wb.active

for row in grand_air:
     ws.append(row)

for cell in ws['1']:
     cell.font = Font(bold=True)

wb.save('air_filters.xlsx')

oil = json.loads(oil)
header = list(oil[0].keys())
grand_oil = [list(item.values()) for item in oil]
grand_oil.insert(0, header)

wb = Workbook()
ws = wb.active

for row in grand_oil:
     ws.append(row)

for cell in ws['1']:
     cell.font = Font(bold=True)

wb.save('oil_filters.xlsx')

